#!/usr/bin/env python3
import os, sys, json, datetime, shutil, traceback
from pathlib import Path
from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
import pandas as pd
import openpyxl
import firebase_admin
from firebase_admin import credentials, firestore

APP_DIR = Path(__file__).parent.resolve()
INPUT_DB = APP_DIR / 'ubpl_Database.xlsx'
if os.environ.get('VERCEL') == '1':
    OUTPUT_DB = Path('/tmp/ubpl_Dashboard_Output.xlsx')
else:
    OUTPUT_DB = APP_DIR / 'ubpl_Dashboard_Output.xlsx'

STATUSES = ['AFC','IFC','IFA','IFR','IDC','DUE']
STATUS_WEIGHTS = {'DUE':0,'IDC':0.15,'IFR':0.4,'IFA':0.65,'IFC':0.8,'AFC':0.95}

app = Flask(__name__)
CORS(app)

# ── Firebase Setup ──
db = None
try:
    fb_cred_json = os.environ.get('FIREBASE_SERVICE_ACCOUNT')
    if fb_cred_json:
        cred_info = json.loads(fb_cred_json)
        firebase_admin.initialize_app(credentials.Certificate(cred_info))
        db = firestore.client()
        print("Firebase initialized from environment variable.")
    else:
        KEY_NAME = 'ramesh-sir-s-dashboard-firebase-adminsdk-fbsvc-5ecf853896.json'
        candidates = [APP_DIR / KEY_NAME, APP_DIR.parent / KEY_NAME]
        for c in candidates:
            if c.exists():
                cred = credentials.Certificate(str(c))
                firebase_admin.initialize_app(cred)
                db = firestore.client()
                print(f"Firebase initialized from local file: {c.name}")
                break
except Exception as e:
    print(f"Firebase initialization failed: {e}")

# ── Dynamic Helpers ──
def to_id(s): return str(s).replace('/', '_').replace('.', '_').strip()

def ensure_setup():
    if not OUTPUT_DB.exists():
        if INPUT_DB.exists():
            try: shutil.copy2(str(INPUT_DB), str(OUTPUT_DB))
            except: pass
        if db:
            try: firestore_pull_all()
            except: pass

@app.before_request
def startup():
    ensure_setup()

@app.errorhandler(Exception)
def handle_err(e):
    return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500

# ── Firestore Sync Logic ──
def firestore_pull_all():
    if not db: return
    try:
        dci_docs = db.collection('deliverables').stream()
        dci_data = [d.to_dict() for d in dci_docs]
        if not dci_data: return
        wb = openpyxl.load_workbook(str(OUTPUT_DB))
        if 'DCI' in wb.sheetnames: del wb['DCI']
        ws = wb.create_sheet('DCI')
        headers = ['Discipline','DocNum','DocDesc','SubFormat','ApprovalCat','Weitage','InitSchedule','LatestRev','LatestStatus','PctWeitage','Remarks']
        for ci, h in enumerate(headers): ws.cell(row=1, column=ci+1, value=h)
        for ri, row in enumerate(dci_data):
            ws.cell(row=ri+2, column=1, value=row.get('discipline',''))
            ws.cell(row=ri+2, column=2, value=row.get('docNum',''))
            ws.cell(row=ri+2, column=3, value=row.get('docDesc',''))
            ws.cell(row=ri+2, column=4, value=row.get('subFormat',''))
            ws.cell(row=ri+2, column=5, value=row.get('approvalCat',''))
            ws.cell(row=ri+2, column=6, value=row.get('weitage',0))
            ws.cell(row=ri+2, column=7, value=row.get('initSchedule',''))
            ws.cell(row=ri+2, column=8, value=row.get('latestRev',''))
            ws.cell(row=ri+2, column=9, value=row.get('latestStatus',''))
            ws.cell(row=ri+2, column=10, value=row.get('pctWeitage',0))
            ws.cell(row=ri+2, column=11, value=row.get('remarks',''))
        wb.save(str(OUTPUT_DB))
        rebuild_summary()
    except Exception as e: print(f"Pull failed: {e}")

# ── Business Logic ──
def read_summary():
    df = pd.read_excel(str(OUTPUT_DB), sheet_name='Summary', engine='openpyxl')
    df = df[df['Discipline'] != 'TOTAL'].copy().fillna(0)
    res = []
    for _, r in df.iterrows():
        d = {'name': str(r['Discipline']), 'AFC': int(r.get('AFC',0)), 'IFC': int(r.get('IFC',0)), 'IFA': int(r.get('IFA',0)), 'IFR': int(r.get('IFR',0)), 'IDC': int(r.get('IDC',0)), 'DUE': int(r.get('DUE',0)), 'total': int(r.get('Total',0)), 'plannedWtg': float(r.get('PlannedWtg',0)), 'earnedWtg': float(r.get('EarnedWtg',0))}
        d['total'] = sum(d[s] for s in STATUSES)
        d['completionPct'] = (d['earnedWtg']/d['plannedWtg']*100) if d['plannedWtg']>0 else 0
        res.append(d)
    return res

def read_dci():
    df = pd.read_excel(str(OUTPUT_DB), sheet_name='DCI', engine='openpyxl').fillna('')
    return [{'discipline':str(r.get('Discipline','')),'docNum':str(r.get('DocNum','')),'docDesc':str(r.get('DocDesc','')), 'subFormat':str(r.get('SubFormat','')),'approvalCat':str(r.get('ApprovalCat','')), 'weitage':float(r.get('Weitage',0)) if r.get('Weitage','')!='' else 0, 'initSchedule':str(r.get('InitSchedule',''))[:10],'latestRev':str(r.get('LatestRev','')), 'latestStatus':str(r.get('LatestStatus','')),'pctWeitage':float(r.get('PctWeitage',0)) if r.get('PctWeitage','')!='' else 0, 'remarks':str(r.get('Remarks',''))} for _,r in df.iterrows()]

def update_status(doc_num, new_status, new_rev=''):
    wb = openpyxl.load_workbook(str(OUTPUT_DB))
    ws = wb['DCI']
    for row in range(2, ws.max_row+1):
        if str(ws.cell(row=row, column=2).value) == str(doc_num):
            ws.cell(row=row, column=9, value=new_status)
            if new_rev: ws.cell(row=row, column=8, value=new_rev)
            wt = ws.cell(row=row, column=6).value or 0
            ws.cell(row=row, column=10, value=float(wt)*STATUS_WEIGHTS.get(new_status,0))
            wb.save(str(OUTPUT_DB)); rebuild_summary()
            if db:
                try: db.collection('deliverables').document(to_id(doc_num)).update({'latestStatus': new_status, 'latestRev': new_rev})
                except: pass
            return True
    return False

def rebuild_summary():
    dci = pd.read_excel(str(OUTPUT_DB), sheet_name='DCI', engine='openpyxl')
    dci.columns = ['Discipline','DocNum','DocDesc','SubFormat','ApprovalCat','Weitage','InitSchedule','LatestRev','LatestStatus','PctWeitage','Remarks']
    ct = pd.crosstab(dci['Discipline'], dci['LatestStatus'])
    for s in STATUSES:
        if s not in ct.columns: ct[s]=0
    wt = dci.groupby('Discipline').agg(PlannedWtg=('Weitage',lambda x:pd.to_numeric(x,errors='coerce').sum()),EarnedWtg=('PctWeitage',lambda x:pd.to_numeric(x,errors='coerce').sum()))
    wb = openpyxl.load_workbook(str(OUTPUT_DB))
    ws = wb['Summary']
    for row in range(2, ws.max_row+1):
        disc = ws.cell(row=row, column=1).value
        if disc and disc!='TOTAL' and disc in ct.index:
            for si,s in enumerate(STATUSES): ws.cell(row=row, column=si+2, value=int(ct.loc[disc].get(s,0)))
            if disc in wt.index: ws.cell(row=row, column=10, value=round(float(wt.loc[disc,'EarnedWtg']),6))
    wb.save(str(OUTPUT_DB))
    if db:
        try:
            summary = {to_id(d['name']): d for d in read_summary()}
            db.collection('project_meta').document('summary').set(summary)
        except: pass

def read_snapshots():
    try:
        df = pd.read_excel(str(OUTPUT_DB), sheet_name='Snapshots', engine='openpyxl').fillna(0)
        snaps = {}
        for _, r in df.iterrows():
            key = str(r.get('SnapshotDate',''))[:10]
            if not key or key == '0': continue
            if key not in snaps: snaps[key] = {'date':key,'week':int(r.get('Week',0)),'year':int(r.get('Year',0)),'disciplines':[]}
            snaps[key]['disciplines'].append({'name':str(r.get('Discipline','')),'AFC':int(r.get('AFC',0)),'IFC':int(r.get('IFC',0)), 'IFA':int(r.get('IFA',0)),'IFR':int(r.get('IFR',0)),'IDC':int(r.get('IDC',0)),'DUE':int(r.get('DUE',0)), 'total':int(r.get('Total',0)),'plannedWtg':float(r.get('PlannedWtg',0)),'earnedWtg':float(r.get('EarnedWtg',0)), 'completionPct':float(r.get('CompletionPct',0))})
        res = []
        for k,v in sorted(snaps.items(), reverse=True):
            t = {'docs':0,'wt':0,'ew':0,'status':{s:0 for s in STATUSES}}
            for d in v['disciplines']:
                t['docs']+=d['total']; t['wt']+=d['plannedWtg']; t['ew']+=d['earnedWtg']
                for s in STATUSES: t['status'][s]+=d[s]
            t['pct']=(t['ew']/t['wt']*100) if t['wt']>0 else 0
            t['approved']=t['status']['AFC']+t['status']['IFC']; t['pending']=t['status']['DUE']; v['totals']=t; res.append(v)
        return res
    except: return []

# ── API Routes ──
@app.route('/api/summary')
def api_summary(): return jsonify(read_summary())
@app.route('/api/dci')
def api_dci():
    data = read_dci()
    disc,status = request.args.get('discipline',''),request.args.get('status','')
    if disc: data=[d for d in data if d['discipline']==disc]
    if status: data=[d for d in data if d['latestStatus']==status]
    return jsonify(data)
@app.route('/api/config')
def api_config():
    df = pd.read_excel(str(OUTPUT_DB), sheet_name='Config', engine='openpyxl')
    return jsonify({str(r['Key']): str(r['Value']) for _, r in df.iterrows()})
@app.route('/api/snapshots')
def api_snapshots(): return jsonify(read_snapshots())
@app.route('/api/update_status', methods=['POST'])
def api_up_status(): d=request.json; return jsonify({'success':update_status(d['docNum'],d['newStatus'],d.get('newRev',''))})
@app.route('/api/db_info')
def api_db_info(): return jsonify({'inputFile': INPUT_DB.name, 'outputFile': OUTPUT_DB.name, 'status': 'online'})
@app.route('/')
def index_api(): return jsonify({"status":"online","detail":"full-featured (firebase included)"})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5500, debug=True)
