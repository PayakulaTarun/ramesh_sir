#!/usr/bin/env python3
"""
UBPL Engineering Dashboard — Standalone Application
====================================================
Reads from INPUT database (ubpl_Database.xlsx) — never modified.
Writes all edits/snapshots to OUTPUT database (ubpl_Dashboard_Output.xlsx).

USAGE:  python ubpl_app.py
Opens browser at http://localhost:5500

Requirements: pip install flask openpyxl pandas
"""
import os, sys, json, webbrowser, threading, datetime, shutil
from pathlib import Path
from flask import Flask, jsonify, request, send_file, Response
import openpyxl
from flask_cors import CORS
import pandas as pd
import firebase_admin
from firebase_admin import credentials, firestore

APP_DIR = Path(__file__).parent.resolve()
ROOT_DIR = APP_DIR.parent
FB_KEY = ROOT_DIR / 'ramesh-sir-s-dashboard-firebase-adminsdk-fbsvc-5ecf853896.json'
INPUT_DB  = APP_DIR / 'ubpl_Database.xlsx'       # READ-ONLY source
OUTPUT_DB = APP_DIR / 'ubpl_Dashboard_Output.xlsx' # All writes go here
PORT = int(os.environ.get('PORT', 5500))
STATUSES = ['AFC','IFC','IFA','IFR','IDC','DUE']
STATUS_WEIGHTS = {'DUE':0,'IDC':0.15,'IFR':0.4,'IFA':0.65,'IFC':0.8,'AFC':0.95}

app = Flask(__name__, static_folder=str(ROOT_DIR / 'dash-frontend' / 'dist'), static_url_path='/')
CORS(app)

# ── Firebase Setup ──
db = None
try:
    fb_cred_json = os.environ.get('FIREBASE_SERVICE_ACCOUNT')
    if fb_cred_json:
        # Load from Environment variable (for Vercel/Render)
        cred_info = json.loads(fb_cred_json)
        firebase_admin.initialize_app(credentials.Certificate(cred_info))
        db = firestore.client()
        print("Firebase initialized from environment variable.")
    else:
        # Fallback to local files (multiple places)
        KEY_NAME = 'ramesh-sir-s-dashboard-firebase-adminsdk-fbsvc-5ecf853896.json'
        candidates = [
            ROOT_DIR / KEY_NAME,
            APP_DIR / KEY_NAME,
            APP_DIR.parent / KEY_NAME
        ]
        key_found = False
        for c in candidates:
            if c.exists():
                cred = credentials.Certificate(str(c))
                firebase_admin.initialize_app(cred)
                db = firestore.client()
                print(f"Firebase initialized from local file: {c.name}")
                key_found = True; break
        if not key_found:
            print("Firebase: No service account key found.")
except Exception as e:
    print(f"Firebase initialization failed: {e}")



def to_id(s):
    """Sanitize string for Firestore Document ID (no slashes)."""
    return str(s).replace('/', '_').replace('.', '_').strip()

def firestore_push_all():
    """Helper to force sync everything from Excel to Firestore."""
    if not db: return
    try:
        # Sync Config
        config = read_config()
        db.collection('project_meta').document('config').set(config)
        # Sync Summary
        summary = {to_id(d['name']): d for d in read_summary()}
        db.collection('project_meta').document('summary').set(summary)
        # Sync DCI
        dci_batch = db.batch()
        for row in read_dci():
            doc_id = to_id(row['docNum'])
            if not doc_id: continue
            doc_ref = db.collection('deliverables').document(doc_id)
            dci_batch.set(doc_ref, row)
        dci_batch.commit()
        print("Firestore: Pushed all data to cloud.")
    except Exception as e:
        print(f"Firestore: Push ALL failed: {e}")

def firestore_pull_all():
    """Helper to restore Excel file from Firestore data."""
    if not db: return
    try:
        # Pull DCI
        dci_docs = db.collection('deliverables').stream()
        dci_data = [d.to_dict() for d in dci_docs]
        if not dci_data: return
        
        wb = openpyxl.load_workbook(str(OUTPUT_DB))
        # Re-write DCI
        if 'DCI' in wb.sheetnames: del wb['DCI']
        ws = wb.create_sheet('DCI')
        headers = ['Discipline','DocNum','DocDesc','SubFormat','ApprovalCat','Weitage','InitSchedule','LatestRev','LatestStatus','PctWeitage','Remarks']
        for ci, h in enumerate(headers): ws.cell(row=1, column=ci+1, value=h)
        for ri, row in enumerate(dci_data):
            ws.cell(row=ri+2, column=1, value=row.get('discipline',''))
            ws.cell(row=ri+2, column=2, value=row.get('docNum',''))
            ws.cell(row=ri+2, column=3, value=row.get('docDesc',''))
            ws.cell(row=ri+2, column=4, value=row.get('subFormat',''))
            ws.cell(row=ri+2, column=5, value=row.get('approvalCat',''))
            ws.cell(row=ri+2, column=6, value=row.get('weitage',0))
            ws.cell(row=ri+2, column=7, value=row.get('initSchedule',''))
            ws.cell(row=ri+2, column=8, value=row.get('latestRev',''))
            ws.cell(row=ri+2, column=9, value=row.get('latestStatus',''))
            ws.cell(row=ri+2, column=10, value=row.get('pctWeitage',0))
            ws.cell(row=ri+2, column=11, value=row.get('remarks',''))
        
        # Pull Config/Summary
        meta_config = db.collection('project_meta').document('config').get()
        if meta_config.exists:
            if 'Config' in wb.sheetnames: del wb['Config']
            ws_c = wb.create_sheet('Config')
            ws_c.cell(row=1, column=1, value='Key')
            ws_c.cell(row=1, column=2, value='Value')
            for ri, (k,v) in enumerate(meta_config.to_dict().items()):
                ws_c.cell(row=ri+2, column=1, value=k)
                ws_c.cell(row=ri+2, column=2, value=v)
        
        # Pull Snapshots (Optional)
        snap_docs = db.collection('snapshots').stream()
        if snap_docs:
            if 'Snapshots' in wb.sheetnames: del wb['Snapshots']
            ws_s = wb.create_sheet('Snapshots')
            snap_headers = ['SnapshotDate','Week','Year','Discipline','AFC','IFC','IFA','IFR','IDC','DUE','Total','PlannedWtg','EarnedWtg','CompletionPct']
            for ci, h in enumerate(snap_headers): ws_s.cell(row=1, column=ci+1, value=h)
            ri = 2
            for s_doc in snap_docs:
                s = s_doc.to_dict()
                for d in s.get('disciplines', []):
                    vals = [s['date'], s['week'], s['year'], d['name'], d['AFC'], d['IFC'], d['IFA'], d['IFR'], d['IDC'], d['DUE'], d['total'], d['plannedWtg'], d['earnedWtg'], d['completionPct']]
                    for ci, v in enumerate(vals): ws_s.cell(row=ri, column=ci+1, value=v)
                    ri += 1

        wb.save(str(OUTPUT_DB))
        rebuild_summary_from_dci()
        print("Firestore: Restored all data to local Excel.")
    except Exception as e:
        print(f"Firestore: Pull failed: {e}")

def ensure_output_db():
    """Create output DB from input if it doesn't exist yet."""
    if not OUTPUT_DB.exists():
        if not INPUT_DB.exists():
            print(f"ERROR: Input database '{INPUT_DB.name}' not found!"); sys.exit(1)
        shutil.copy2(str(INPUT_DB), str(OUTPUT_DB))
        print(f"Created output database: {OUTPUT_DB.name}")

def reimport_from_input():
    """Re-import DCI + Summary + Config from input DB into output DB, preserving snapshots."""
    if not INPUT_DB.exists(): return False
    inp = openpyxl.load_workbook(str(INPUT_DB))
    out = openpyxl.load_workbook(str(OUTPUT_DB))
    for sheet_name in ['DCI','Summary','Config']:
        if sheet_name in inp.sheetnames:
            if sheet_name in out.sheetnames:
                del out[sheet_name]
            src = inp[sheet_name]
            dst = out.create_sheet(sheet_name)
            for row in src.iter_rows(values_only=False):
                for cell in row:
                    dst.cell(row=cell.row, column=cell.column, value=cell.value)
    out.save(str(OUTPUT_DB))
    return True

# ── Read helpers (always read from OUTPUT_DB) ──
def read_summary():
    df = pd.read_excel(str(OUTPUT_DB), sheet_name='Summary', engine='openpyxl')
    df = df[df['Discipline'] != 'TOTAL'].copy().fillna(0)
    result = []
    for _, r in df.iterrows():
        d = {'name': str(r['Discipline']),
             'AFC': int(r.get('AFC',0)), 'IFC': int(r.get('IFC',0)),
             'IFA': int(r.get('IFA',0)), 'IFR': int(r.get('IFR',0)),
             'IDC': int(r.get('IDC',0)), 'DUE': int(r.get('DUE',0)),
             'total': int(r.get('Total',0)),
             'plannedWtg': float(r.get('PlannedWtg',0)),
             'earnedWtg': float(r.get('EarnedWtg',0))}
        d['total'] = sum(d[s] for s in STATUSES)
        d['completionPct'] = (d['earnedWtg']/d['plannedWtg']*100) if d['plannedWtg']>0 else 0
        result.append(d)
    return result

def read_dci():
    df = pd.read_excel(str(OUTPUT_DB), sheet_name='DCI', engine='openpyxl').fillna('')
    return [{'discipline':str(r.get('Discipline','')),'docNum':str(r.get('DocNum','')),'docDesc':str(r.get('DocDesc','')),
             'subFormat':str(r.get('SubFormat','')),'approvalCat':str(r.get('ApprovalCat','')),
             'weitage':float(r.get('Weitage',0)) if r.get('Weitage','')!='' else 0,
             'initSchedule':str(r.get('InitSchedule',''))[:10],'latestRev':str(r.get('LatestRev','')),
             'latestStatus':str(r.get('LatestStatus','')),'pctWeitage':float(r.get('PctWeitage',0)) if r.get('PctWeitage','')!='' else 0,
             'remarks':str(r.get('Remarks',''))} for _,r in df.iterrows()]

def read_config():
    df = pd.read_excel(str(OUTPUT_DB), sheet_name='Config', engine='openpyxl')
    return {str(r['Key']): str(r['Value']) for _, r in df.iterrows()}

def read_snapshots():
    try:
        df = pd.read_excel(str(OUTPUT_DB), sheet_name='Snapshots', engine='openpyxl')
        if len(df)==0: return []
        df = df.fillna(0)
        snaps = {}
        for _, r in df.iterrows():
            key = str(r.get('SnapshotDate',''))[:10]
            if key not in snaps:
                snaps[key] = {'date':key,'week':int(r.get('Week',0)),'year':int(r.get('Year',0)),'disciplines':[]}
            snaps[key]['disciplines'].append({'name':str(r.get('Discipline','')),'AFC':int(r.get('AFC',0)),'IFC':int(r.get('IFC',0)),
                'IFA':int(r.get('IFA',0)),'IFR':int(r.get('IFR',0)),'IDC':int(r.get('IDC',0)),'DUE':int(r.get('DUE',0)),
                'total':int(r.get('Total',0)),'plannedWtg':float(r.get('PlannedWtg',0)),'earnedWtg':float(r.get('EarnedWtg',0)),
                'completionPct':float(r.get('CompletionPct',0))})
        result = []
        for k,v in sorted(snaps.items(), reverse=True):
            t = {'docs':0,'wt':0,'ew':0,'status':{s:0 for s in STATUSES}}
            for d in v['disciplines']:
                t['docs']+=d['total']; t['wt']+=d['plannedWtg']; t['ew']+=d['earnedWtg']
                for s in STATUSES: t['status'][s]+=d[s]
            t['pct']=(t['ew']/t['wt']*100) if t['wt']>0 else 0
            t['approved']=t['status']['AFC']+t['status']['IFC']; t['pending']=t['status']['DUE']
            v['totals']=t; result.append(v)
        return result
    except: return []

# ── Write helpers (always write to OUTPUT_DB) ──
def update_summary_cell(discipline, status, value):
    wb = openpyxl.load_workbook(str(OUTPUT_DB))
    ws = wb['Summary']
    col_map = {s:i+2 for i,s in enumerate(STATUSES)}
    if status not in col_map: return False
    for row in range(2, ws.max_row+1):
        if ws.cell(row=row, column=1).value == discipline:
            ws.cell(row=row, column=col_map[status], value=int(value))
            ws.cell(row=row, column=8).value = f'=SUM(B{row}:G{row})'
            break
    wb.save(str(OUTPUT_DB)); return True

def update_dci_status(doc_num, new_status, new_rev=''):
    wb = openpyxl.load_workbook(str(OUTPUT_DB))
    ws = wb['DCI']
    for row in range(2, ws.max_row+1):
        if ws.cell(row=row, column=2).value == doc_num:
            ws.cell(row=row, column=9, value=new_status)
            if new_rev: ws.cell(row=row, column=8, value=new_rev)
            wt = ws.cell(row=row, column=6).value or 0
            ws.cell(row=row, column=10, value=float(wt)*STATUS_WEIGHTS.get(new_status,0))
            wb.save(str(OUTPUT_DB)); 
            rebuild_summary_from_dci(); 
            if db: 
                try: db.collection('deliverables').document(to_id(doc_num)).update({'latestStatus': new_status, 'latestRev': new_rev})
                except Exception as e: print(f"Firestore: Update failed: {e}")
            return True
    return False

def add_new_dci(data):
    try:
        wb = openpyxl.load_workbook(str(OUTPUT_DB))
        ws = wb['DCI']
        nr = ws.max_row + 1
        # Columns: Discipline, DocNum, DocDesc, SubFormat, ApprovalCat, Weitage, InitSchedule, LatestRev, LatestStatus, PctWeitage, Remarks
        wt = float(data.get('weitage', 0))
        status = data.get('latestStatus', 'DUE')
        pct_wt = wt * STATUS_WEIGHTS.get(status, 0)
        
        row_data = [
            data.get('discipline', ''),
            data.get('docNum', ''),
            data.get('docDesc', ''),
            data.get('subFormat', ''),
            data.get('approvalCat', ''),
            wt,
            data.get('initSchedule', ''),
            data.get('latestRev', ''),
            status,
            pct_wt,
            data.get('remarks', '')
        ]
        for ci, v in enumerate(row_data):
            ws.cell(row=nr, column=ci+1, value=v)
        wb.save(str(OUTPUT_DB))
        if db:
            try:
                db.collection('deliverables').document(to_id(data.get('docNum'))).set({
                    'discipline': data.get('discipline',''), 'docNum': data.get('docNum',''), 'docDesc': data.get('docDesc',''),
                    'subFormat': data.get('subFormat',''), 'approvalCat': data.get('approvalCat',''), 'weitage': wt,
                    'initSchedule': data.get('initSchedule',''), 'latestRev': data.get('latestRev',''), 'latestStatus': status,
                    'pctWeitage': pct_wt, 'remarks': data.get('remarks','')
                })
            except Exception as e: print(f"Firestore: Add failed: {e}")
        return True
    except Exception as e:
        print(f"Error adding DCI: {e}")
        return False

def delete_dci(doc_num):
    try:
        wb = openpyxl.load_workbook(str(OUTPUT_DB))
        ws = wb['DCI']
        for row in range(2, ws.max_row+1):
            if ws.cell(row=row, column=2).value == doc_num:
                ws.delete_rows(row)
                wb.save(str(OUTPUT_DB))
                if db: 
                    try: db.collection('deliverables').document(to_id(doc_num)).delete()
                    except Exception as e: print(f"Firestore: Delete failed: {e}")
                return True
        return False
    except Exception as e:
        print(f"Error deleting DCI: {e}")
        return False

def rebuild_summary_from_dci():
    dci = pd.read_excel(str(OUTPUT_DB), sheet_name='DCI', engine='openpyxl')
    dci.columns = ['Discipline','DocNum','DocDesc','SubFormat','ApprovalCat','Weitage','InitSchedule','LatestRev','LatestStatus','PctWeitage','Remarks']
    ct = pd.crosstab(dci['Discipline'], dci['LatestStatus'])
    for s in STATUSES:
        if s not in ct.columns: ct[s]=0
    wt = dci.groupby('Discipline').agg(PlannedWtg=('Weitage',lambda x:pd.to_numeric(x,errors='coerce').sum()),EarnedWtg=('PctWeitage',lambda x:pd.to_numeric(x,errors='coerce').sum()))
    wb = openpyxl.load_workbook(str(OUTPUT_DB))
    ws = wb['Summary']
    for row in range(2, ws.max_row+1):
        disc = ws.cell(row=row, column=1).value
        if disc and disc!='TOTAL' and disc in ct.index:
            for si,s in enumerate(STATUSES): ws.cell(row=row, column=si+2, value=int(ct.loc[disc].get(s,0)))
            if disc in wt.index: ws.cell(row=row, column=10, value=round(float(wt.loc[disc,'EarnedWtg']),6))
    wb.save(str(OUTPUT_DB))
    if db:
        try:
            summary = {to_id(d['name']): d for d in read_summary()}
            db.collection('project_meta').document('summary').set(summary)
        except Exception as e: print(f"Firestore: Summary Sync failed: {e}")

def save_snapshot(date_str):
    summary = read_summary()
    dt = datetime.datetime.strptime(date_str, '%Y-%m-%d')
    week, year = dt.isocalendar()[1], dt.year
    wb = openpyxl.load_workbook(str(OUTPUT_DB))
    ws = wb['Snapshots']
    nr = ws.max_row+1
    if ws.cell(row=1,column=1).value and not ws.cell(row=2,column=1).value: nr=2
    for d in summary:
        pct=(d['earnedWtg']/d['plannedWtg']*100) if d['plannedWtg']>0 else 0
        for ci,v in enumerate([date_str,week,year,d['name'],d['AFC'],d['IFC'],d['IFA'],d['IFR'],d['IDC'],d['DUE'],d['total'],d['plannedWtg'],d['earnedWtg'],round(pct,2)]):
            ws.cell(row=nr, column=ci+1, value=v)
        nr+=1
    wb.save(str(OUTPUT_DB))
    if db:
        try: db.collection('snapshots').document(to_id(date_str)).set(save_snapshot_firestore(date_str, week, year, summary))
        except Exception as e: print(f"Firestore: Snapshot failed: {e}")
    return {'week':week,'year':year,'date':date_str,'count':len(summary)}

def save_snapshot_firestore(date_str, week, year, summary):
    return {'date': date_str, 'week': week, 'year': year, 'disciplines': summary}

# ── API Routes ──
@app.route('/api/summary')
def api_summary(): return jsonify(read_summary())

@app.route('/api/dci')
def api_dci():
    data = read_dci()
    disc,status,cat = request.args.get('discipline',''),request.args.get('status',''),request.args.get('category','')
    if disc: data=[d for d in data if d['discipline']==disc]
    if status: data=[d for d in data if d['latestStatus']==status]
    if cat: data=[d for d in data if d['approvalCat']==cat]
    return jsonify(data)

@app.route('/api/config')
def api_config(): return jsonify(read_config())

@app.route('/api/snapshots')
def api_snapshots(): return jsonify(read_snapshots())

@app.route('/api/update_status', methods=['POST'])
def api_update_status():
    d=request.json; return jsonify({'success':update_dci_status(d['docNum'],d['newStatus'],d.get('newRev',''))})

@app.route('/api/update_summary', methods=['POST'])
def api_update_summary():
    d=request.json; return jsonify({'success':update_summary_cell(d['discipline'],d['status'],d['value'])})

@app.route('/api/save_snapshot', methods=['POST'])
def api_save_snapshot(): return jsonify(save_snapshot(request.json['date']))

@app.route('/api/add_dci', methods=['POST'])
def api_add_dci():
    success = add_new_dci(request.json)
    if success: rebuild_summary_from_dci()
    return jsonify({'success': success})

@app.route('/api/delete_dci', methods=['POST'])
def api_delete_dci():
    success = delete_dci(request.json['docNum'])
    if success: rebuild_summary_from_dci()
    return jsonify({'success': success})

@app.route('/api/refresh_summary', methods=['POST'])
def api_refresh(): rebuild_summary_from_dci(); return jsonify({'success':True})

@app.route('/api/reimport', methods=['POST'])
def api_reimport():
    ok = reimport_from_input()
    if ok and db: firestore_push_all()
    return jsonify({'success':ok,'message':'Re-imported from input DB + Synced to Firestore' if ok else 'Input DB not found'})

@app.route('/api/download_output')
def api_download_output(): return send_file(str(OUTPUT_DB), as_attachment=True, download_name='ubpl_Dashboard_Output.xlsx')

@app.route('/api/download_input')
def api_download_input():
    if INPUT_DB.exists(): return send_file(str(INPUT_DB), as_attachment=True, download_name='ubpl_Database.xlsx')
    return jsonify({'error':'Input DB not found'}), 404

@app.route('/api/db_info')
def api_db_info():
    return jsonify({
        'inputFile': INPUT_DB.name, 'inputExists': INPUT_DB.exists(),
        'inputModified': datetime.datetime.fromtimestamp(INPUT_DB.stat().st_mtime).isoformat() if INPUT_DB.exists() else None,
        'outputFile': OUTPUT_DB.name, 'outputExists': OUTPUT_DB.exists(),
        'outputModified': datetime.datetime.fromtimestamp(OUTPUT_DB.stat().st_mtime).isoformat() if OUTPUT_DB.exists() else None,
    })

@app.route('/favicon.ico')
def favicon():
    return '', 204

@app.route('/')
def index_api(): 
    return jsonify({
        "status": "online",
        "message": "UBPL Engineering Dashboard API is running",
        "version": "1.0.0",
        "database": {
            "input": INPUT_DB.name,
            "output": OUTPUT_DB.name
        }
    })

# ── Startup Logic ──
ensure_output_db()
if db:
    try:
        firestore_pull_all() 
    except Exception as e:
        print(f"Startup Firestore pull failed: {e}")

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5500))
    print(f"Starting server on port {port}...")
    app.run(host='0.0.0.0', port=port, debug=False)
