#!/usr/bin/env python3
import os, sys, json, datetime, shutil
from pathlib import Path
from flask import Flask, jsonify, request, send_file
import openpyxl
from flask_cors import CORS
from supabase import create_client, Client

# Vercel Paths
APP_DIR = Path(__file__).parent.resolve()
INPUT_DB = APP_DIR / 'ubpl_Database.xlsx'
if os.environ.get('VERCEL') == '1':
    OUTPUT_DB = Path('/tmp/ubpl_Dashboard_Output.xlsx')
else:
    OUTPUT_DB = APP_DIR / 'ubpl_Dashboard_Output.xlsx'

PORT = int(os.environ.get('PORT', 5500))
STATUSES = ['AFC','IFC','IFA','IFR','IDC','DUE']

app = Flask(__name__)
CORS(app)

def ensure_output_db():
    if not OUTPUT_DB.exists() and INPUT_DB.exists():
        shutil.copy2(str(INPUT_DB), str(OUTPUT_DB))

@app.route('/api/summary')
def api_sum():
    ensure_output_db()
    try:
        wb = openpyxl.load_workbook(str(OUTPUT_DB), data_only=True)
        ws = wb['Summary']
        data = []; headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0] or str(row[0]) == 'TOTAL': continue
            d = {headers[i]: v for i, v in enumerate(row) if i < len(headers)}
            data.append({
                'name': str(d.get('Discipline', '')),
                'AFC': int(d.get('AFC', 0) or 0), 'IFC': int(d.get('IFC', 0) or 0),
                'IFA': int(d.get('IFA', 0) or 0), 'IFR': int(d.get('IFR', 0) or 0),
                'IDC': int(d.get('IDC', 0) or 0), 'DUE': int(d.get('DUE', 0) or 0),
                'plannedWtg': float(d.get('PlannedWtg', 0) or 0),
                'earnedWtg': float(d.get('EarnedWtg', 0) or 0),
                'total': 0, 'completionPct': 0
            })
        return jsonify(data)
    except: return jsonify([])

@app.route('/api/config')
def api_conf():
    ensure_output_db()
    try:
        wb = openpyxl.load_workbook(str(OUTPUT_DB), data_only=True)
        ws = wb['Config']
        return jsonify({str(r[0]): str(r[1]) for r in ws.iter_rows(min_row=2, values_only=True) if r[0]})
    except: return jsonify({})

@app.route('/api/snapshots')
def api_snaps(): return jsonify([])

@app.route('/api/db_info')
def api_db():
    return jsonify({
        'inputFile': INPUT_DB.name if INPUT_DB.exists() else 'Missing',
        'outputFile': OUTPUT_DB.name,
        'status': 'proxy-online'
    })

@app.route('/')
def ide_fe(): return jsonify({"status":"online"})

ensure_output_db()

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=PORT, debug=True)
